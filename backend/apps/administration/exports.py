"""Excel export views for reports."""
from io import BytesIO
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.utils.html import escape
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import HasRolePermission
from apps.students.models import Student
from apps.administration.models import Invoice, Payment
from apps.flight_training.models import FlightLesson
from apps.core.models import AuditLog
from openpyxl import Workbook
from apps.administration.export_utils import style_header, xlsx_response


class ExportUsersView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'users.export'

    def get(self, request):
        UserModel = get_user_model()
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        style_header(ws, ["Email", "Full Name", "Username", "Role", "Status", "Active"])
        for i, u in enumerate(UserModel.objects.all(), 2):
            ws.cell(row=i, column=1, value=u.email)
            ws.cell(row=i, column=2, value=u.get_full_name() or u.email)
            ws.cell(row=i, column=3, value=u.username)
            ws.cell(row=i, column=4, value=u.role)
            ws.cell(row=i, column=5, value=u.status)
            ws.cell(row=i, column=6, value=u.is_active)
        buf = BytesIO()
        wb.save(buf)
        return HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=users.xlsx"})


class ExportPaymentsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'invoicing.export'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        style_header(ws, ["Student", "Invoice", "Amount", "Date", "Method", "Status"])
        for i, p in enumerate(Payment.objects.select_related('student', 'invoice').all(), 2):
            ws.cell(row=i, column=1, value=p.student.full_name)
            ws.cell(row=i, column=2, value=p.invoice.invoice_number if p.invoice else "")
            ws.cell(row=i, column=3, value=float(p.amount))
            ws.cell(row=i, column=4, value=str(p.paid_at)[:10] if p.paid_at else "")
            ws.cell(row=i, column=5, value=p.method or "")
            ws.cell(row=i, column=6, value=p.invoice.status if p.invoice else "")
        buf = BytesIO()
        wb.save(buf)
        return HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=payments.xlsx"})


class ExportStudentsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'students.export'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        style_header(ws, ["Student Number", "First Name", "Last Name", "Program", "Status", "Enrollment Date"])
        for i, s in enumerate(Student.objects.all(), 2):
            ws.cell(row=i, column=1, value=s.student_number)
            ws.cell(row=i, column=2, value=s.first_name)
            ws.cell(row=i, column=3, value=s.last_name)
            ws.cell(row=i, column=4, value=s.program)
            ws.cell(row=i, column=5, value=s.status)
            ws.cell(row=i, column=6, value=str(s.enrollment_date))
        return xlsx_response(wb, "students.xlsx")


class ExportInvoicesView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'invoices.export'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        style_header(ws, ["Invoice #", "Student", "Type", "Amount", "Currency", "Status", "Issued", "Due"])
        for i, inv in enumerate(Invoice.objects.select_related('student').all(), 2):
            ws.cell(row=i, column=1, value=inv.invoice_number)
            ws.cell(row=i, column=2, value=inv.student.full_name)
            ws.cell(row=i, column=3, value=inv.type or "")
            ws.cell(row=i, column=4, value=float(inv.amount))
            ws.cell(row=i, column=5, value=inv.currency)
            ws.cell(row=i, column=6, value=inv.status)
            ws.cell(row=i, column=7, value=str(inv.issued_at)[:10] if inv.issued_at else "")
            ws.cell(row=i, column=8, value=str(inv.due_at)[:10] if inv.due_at else "")
        return xlsx_response(wb, "invoices.xlsx")


class ExportFlightsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'flights.export'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Flights"
        style_header(ws, ["Date", "Student", "Instructor", "Aircraft", "Duration", "Status", "Grade", "Result"])
        for i, f in enumerate(FlightLesson.objects.select_related('student', 'instructor', 'aircraft').all(), 2):
            ws.cell(row=i, column=1, value=str(f.scheduled_date))
            ws.cell(row=i, column=2, value=f.student.full_name)
            ws.cell(row=i, column=3, value=f"{f.instructor.first_name} {f.instructor.last_name}")
            ws.cell(row=i, column=4, value=f.aircraft.registration)
            ws.cell(row=i, column=5, value=float(f.flight_duration) if f.flight_duration else 0)
            ws.cell(row=i, column=6, value=f.status)
            ws.cell(row=i, column=7, value=float(f.grade) if f.grade else "")
            ws.cell(row=i, column=8, value=f.result or "")
        return xlsx_response(wb, "flights.xlsx")


class FlightsPdfView(APIView):
    """Generate a PDF report listing all flights."""
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'flight_training.view'

    def get(self, request):
        flights = FlightLesson.objects.select_related('student', 'instructor', 'aircraft').all()

        rows = ""
        for f in flights:
            status_color = {
                'scheduled': '#3b82f6',
                'in_progress': '#f59e0b',
                'completed': '#22c55e',
                'cancelled': '#ef4444',
                'postponed': '#8b5cf6',
            }.get(f.status, '#6b7280')
            rows += (
                f"<tr>"
                f"<td>{str(f.scheduled_date) if f.scheduled_date else ''}</td>"
                f"<td>{escape(f.student.full_name)}</td>"
                f"<td>{escape(f'{f.instructor.first_name} {f.instructor.last_name}')}</td>"
                f"<td>{escape(f.aircraft.registration)}</td>"
                f"<td>{float(f.flight_duration) if f.flight_duration else 0}</td>"
                f"<td style='color:{status_color};font-weight:bold'>{escape(f.status or '')}</td>"
                f"<td>{float(f.grade) if f.grade else ''}</td>"
                f"<td>{escape(f.result or '')}</td>"
                f"</tr>"
            )

        total = flights.count()
        completed = flights.filter(status='completed').count()
        scheduled = flights.filter(status='scheduled').count()

        html = f"""<html><head><meta charset="utf-8"><style>
        @page {{ size: A4 landscape; margin: 1.5cm; }}
        body {{ font-family: sans-serif; }}
        .header {{ border-bottom: 2px solid #c4943c; padding-bottom: 10px; margin-bottom: 20px; }}
        .logo {{ font-size: 24px; color: #c4943c; font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 10px; }}
        th {{ background: #0a1628; color: #c4943c; padding: 8px; text-align: left; font-size: 10px; }}
        td {{ padding: 6px 8px; border-bottom: 1px solid #e5e7eb; }}
        .stats {{ display: flex; gap: 20px; margin: 15px 0; }}
        .stat {{ background: #f3f4f6; padding: 10px 20px; border-radius: 8px; }}
        </style></head><body>
        <div class="header"><div class="logo">MAA</div><div>Masterly Air Academy</div><h2>Flights Report</h2></div>
        <div class="stats">
          <div class="stat"><strong>Total:</strong> {total} flights</div>
          <div class="stat"><strong>Scheduled:</strong> {scheduled}</div>
          <div class="stat"><strong>Completed:</strong> {completed}</div>
        </div>
        <table><tr><th>Date</th><th>Student</th><th>Instructor</th><th>Aircraft</th><th>Duration (h)</th><th>Status</th><th>Grade</th><th>Result</th></tr>{rows}</table>
        </body></html>"""

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = 'attachment; filename="flights.pdf"'
            return resp
        except ImportError:
            return HttpResponse("PDF generation not available", status=501)


class ExportAuditLogsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'audit.export'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "AuditLogs"
        style_header(ws, ["Date", "User", "Action", "Entity", "Entity ID", "IP Address", "User Agent"])
        for i, log in enumerate(AuditLog.objects.select_related('user').all(), 2):
            ws.cell(row=i, column=1, value=str(log.created_at)[:19] if log.created_at else "")
            ws.cell(row=i, column=2, value=log.user.email if log.user else "")
            ws.cell(row=i, column=3, value=log.action)
            ws.cell(row=i, column=4, value=log.entity)
            ws.cell(row=i, column=5, value=str(log.entity_id) if log.entity_id else "")
            ws.cell(row=i, column=6, value=log.ip_address or "")
            ws.cell(row=i, column=7, value=str(log.user_agent or "")[:200])
        return xlsx_response(wb, "audit_logs.xlsx")


class ExportCertificatesView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'exams.view'

    def get(self, request):
        from apps.exams.models import Certificate
        from apps.students.models import Student

        # Students can only export their own certificates
        try:
            student = Student.objects.get(user=request.user)
            certificates = Certificate.objects.filter(student=student).select_related('student')
        except Student.DoesNotExist:
            # Admins/staff export all certificates
            certificates = Certificate.objects.select_related('student').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Certificates"
        style_header(ws, ["Certificate #", "Student", "Type", "Title", "Program", "Issue Date", "Expiry Date", "Status"])
        for i, c in enumerate(certificates, 2):
            ws.cell(row=i, column=1, value=c.certificate_number or "")
            ws.cell(row=i, column=2, value=c.student.full_name)
            ws.cell(row=i, column=3, value=c.type or "")
            ws.cell(row=i, column=4, value=c.title or "")
            ws.cell(row=i, column=5, value=c.program or "")
            ws.cell(row=i, column=6, value=str(c.issue_date) if c.issue_date else "")
            ws.cell(row=i, column=7, value=str(c.expiry_date) if c.expiry_date else "")
            ws.cell(row=i, column=8, value=c.status or "")
        return xlsx_response(wb, "certificates.xlsx")


class ExportCoursesView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'ground_training.view'

    def get(self, request):
        from apps.ground_training.models import Course
        courses = Course.objects.select_related('subject', 'instructor', 'room').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Courses"
        style_header(ws, ["Title", "Subject", "Instructor", "Date", "Start", "End", "Room", "Status"])
        for i, c in enumerate(courses, 2):
            ws.cell(row=i, column=1, value=c.title or "")
            ws.cell(row=i, column=2, value=c.subject.code if c.subject else "")
            ws.cell(row=i, column=3, value=f'{c.instructor.first_name} {c.instructor.last_name}' if c.instructor else "")
            ws.cell(row=i, column=4, value=str(c.scheduled_date) if c.scheduled_date else "")
            ws.cell(row=i, column=5, value=str(c.start_time) if c.start_time else "")
            ws.cell(row=i, column=6, value=str(c.end_time) if c.end_time else "")
            ws.cell(row=i, column=7, value=c.room.name if c.room else "")
            ws.cell(row=i, column=8, value=c.status or "")
        return xlsx_response(wb, "courses.xlsx")


class CoursesPdfView(APIView):
    """Generate a PDF report listing all courses."""
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'ground_training.view'

    def get(self, request):
        from apps.ground_training.models import Course
        courses = Course.objects.select_related('subject', 'instructor', 'room').all()

        rows = ""
        for c in courses:
            status_color = {
                'scheduled': '#3b82f6',
                'in_progress': '#f59e0b',
                'completed': '#22c55e',
                'cancelled': '#ef4444',
            }.get(c.status, '#6b7280')
            rows += (
                f"<tr>"
                f"<td>{escape(c.title or '')}</td>"
                f"<td>{escape(c.subject.code) if c.subject else ''}</td>"
                f"<td>{escape(f'{c.instructor.first_name} {c.instructor.last_name}') if c.instructor else ''}</td>"
                f"<td>{str(c.scheduled_date) if c.scheduled_date else ''}</td>"
                f"<td>{str(c.start_time)[:5] if c.start_time else ''} - {str(c.end_time)[:5] if c.end_time else ''}</td>"
                f"<td>{escape(c.room.name) if c.room else ''}</td>"
                f"<td style='color:{status_color};font-weight:bold'>{escape(c.status or '')}</td>"
                f"</tr>"
            )

        total = courses.count()
        scheduled = courses.filter(status='scheduled').count()
        completed = courses.filter(status='completed').count()

        html = f"""<html><head><meta charset="utf-8"><style>
        @page {{ size: A4 landscape; margin: 1.5cm; }}
        body {{ font-family: sans-serif; }}
        .header {{ border-bottom: 2px solid #c4943c; padding-bottom: 10px; margin-bottom: 20px; }}
        .logo {{ font-size: 24px; color: #c4943c; font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 10px; }}
        th {{ background: #0a1628; color: #c4943c; padding: 8px; text-align: left; font-size: 10px; }}
        td {{ padding: 6px 8px; border-bottom: 1px solid #e5e7eb; }}
        .stats {{ display: flex; gap: 20px; margin: 15px 0; }}
        .stat {{ background: #f3f4f6; padding: 10px 20px; border-radius: 8px; }}
        </style></head><body>
        <div class="header"><div class="logo">MAA</div><div>Masterly Air Academy</div><h2>Courses Report</h2></div>
        <div class="stats">
          <div class="stat"><strong>Total:</strong> {total} courses</div>
          <div class="stat"><strong>Scheduled:</strong> {scheduled}</div>
          <div class="stat"><strong>Completed:</strong> {completed}</div>
        </div>
        <table><tr><th>Title</th><th>Subject</th><th>Instructor</th><th>Date</th><th>Time</th><th>Room</th><th>Status</th></tr>{rows}</table>
        </body></html>"""

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = 'attachment; filename="courses.pdf"'
            return resp
        except ImportError:
            return HttpResponse("PDF generation not available", status=501)


class ExportExamsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'exams.view'

    def get(self, request):
        from django.db.models import Count, Q
        from apps.exams.models import Exam
        exams = Exam.objects.select_related('subject').annotate(
            attempt_count=Count('attempts', distinct=True),
            passed_count=Count('attempts', filter=Q(attempts__is_passed=True), distinct=True),
        ).order_by('code')

        wb = Workbook()
        ws = wb.active
        ws.title = "Exams"
        style_header(ws, ["ID", "Code", "Title", "Subject", "Program", "Type", "Status", "Attempts", "Pass Rate (%)"])
        for i, e in enumerate(exams, 2):
            pass_rate = round((e.passed_count / e.attempt_count * 100), 1) if e.attempt_count else 0
            ws.cell(row=i, column=1, value=str(e.id))
            ws.cell(row=i, column=2, value=e.code or "")
            ws.cell(row=i, column=3, value=e.title or "")
            ws.cell(row=i, column=4, value=e.subject.code if e.subject else "")
            ws.cell(row=i, column=5, value=e.program or "")
            ws.cell(row=i, column=6, value=e.type or "")
            ws.cell(row=i, column=7, value=e.status or "")
            ws.cell(row=i, column=8, value=e.attempt_count)
            ws.cell(row=i, column=9, value=pass_rate)
        return xlsx_response(wb, "exams.xlsx")
