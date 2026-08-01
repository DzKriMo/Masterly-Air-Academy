"""Shared helpers for Excel export views."""
from io import BytesIO
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment


def style_header(ws, headers, row=1):
    header_fill = PatternFill(start_color="0a1628", end_color="0a1628", fill_type="solid")
    header_font = Font(color="c4943c", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def xlsx_response(workbook, filename):
    buf = BytesIO()
    workbook.save(buf)
    return HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
