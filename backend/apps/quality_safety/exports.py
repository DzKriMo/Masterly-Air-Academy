"""Excel export views for Quality & Safety."""
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import HasRolePermission
from openpyxl import Workbook
from apps.administration.export_utils import style_header, xlsx_response


class ExportAuditsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'quality.view'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Audits"
        style_header(ws, ["Title", "Type", "Status", "Scheduled", "Completed", "Lead Auditor", "NCR Count"])
        from .models import Audit
        for i, a in enumerate(Audit.objects.select_related('lead_auditor').all(), 2):
            ws.cell(row=i, column=1, value=a.title or "")
            ws.cell(row=i, column=2, value=a.type or "")
            ws.cell(row=i, column=3, value=a.status or "")
            ws.cell(row=i, column=4, value=str(a.scheduled_date)[:10] if a.scheduled_date else "")
            ws.cell(row=i, column=5, value=str(a.completed_date)[:10] if a.completed_date else "")
            ws.cell(row=i, column=6, value=a.lead_auditor.email if a.lead_auditor else "")
            ws.cell(row=i, column=7, value=a.non_conformities.count())
        return xlsx_response(wb, "audits.xlsx")


class ExportNCRsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'quality.view'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "NCRs"
        style_header(ws, ["NCR #", "Title", "Audit", "Severity", "Status", "Responsible", "Due Date"])
        from .models import NonConformity
        for i, n in enumerate(NonConformity.objects.select_related('audit', 'responsible').all(), 2):
            ws.cell(row=i, column=1, value=n.ncr_number or "")
            ws.cell(row=i, column=2, value=n.title or "")
            ws.cell(row=i, column=3, value=n.audit.title if n.audit else "")
            ws.cell(row=i, column=4, value=n.severity or "")
            ws.cell(row=i, column=5, value=n.status or "")
            ws.cell(row=i, column=6, value=n.responsible.email if n.responsible else "")
            ws.cell(row=i, column=7, value=str(n.due_date)[:10] if n.due_date else "")
        return xlsx_response(wb, "non_conformities.xlsx")


class ExportCAPAsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'quality.view'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "CAPAs"
        style_header(ws, ["CAPA #", "Title", "Type", "Related NCR", "Status", "Responsible", "Due Date", "Validation"])
        from .models import CAPA
        for i, c in enumerate(CAPA.objects.select_related('non_conformity', 'responsible').all(), 2):
            ws.cell(row=i, column=1, value=c.capa_number or "")
            ws.cell(row=i, column=2, value=c.title or "")
            ws.cell(row=i, column=3, value=c.type or "")
            ws.cell(row=i, column=4, value=c.non_conformity.title if c.non_conformity else "")
            ws.cell(row=i, column=5, value=c.status or "")
            ws.cell(row=i, column=6, value=c.responsible.email if c.responsible else "")
            ws.cell(row=i, column=7, value=str(c.due_date)[:10] if c.due_date else "")
            ws.cell(row=i, column=8, value=str(c.validation_date)[:10] if c.validation_date else "")
        return xlsx_response(wb, "capas.xlsx")


class ExportSafetyEventsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'safety.view'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "SafetyEvents"
        style_header(ws, ["Title", "Type", "Status", "Reported By", "Confidential", "Created"])
        from .models import SafetyEvent
        for i, e in enumerate(SafetyEvent.objects.select_related('reported_by').all(), 2):
            ws.cell(row=i, column=1, value=e.title or "")
            ws.cell(row=i, column=2, value=e.type or "")
            ws.cell(row=i, column=3, value=e.status or "")
            ws.cell(row=i, column=4, value=e.reported_by.email if e.reported_by else "")
            ws.cell(row=i, column=5, value="Yes" if e.confidential else "No")
            ws.cell(row=i, column=6, value=str(e.created_at)[:10] if e.created_at else "")
        return xlsx_response(wb, "safety_events.xlsx")


class ExportRiskAssessmentsView(APIView):
    permission_classes = [IsAuthenticated, HasRolePermission]
    required_permission = 'safety.view'

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "RiskAssessments"
        style_header(ws, ["Hazard", "Risk Level", "Probability", "Severity", "Status", "Mitigation", "Responsible", "Review Date"])
        from .models import RiskAssessment
        for i, r in enumerate(RiskAssessment.objects.all(), 2):
            ws.cell(row=i, column=1, value=r.hazard or "")
            ws.cell(row=i, column=2, value=r.risk_level)
            ws.cell(row=i, column=3, value=r.probability)
            ws.cell(row=i, column=4, value=r.severity)
            ws.cell(row=i, column=5, value=r.status or "")
            ws.cell(row=i, column=6, value=r.mitigation_measures or "")
            ws.cell(row=i, column=7, value=r.responsible.email if r.responsible else "")
            ws.cell(row=i, column=8, value=str(r.reeval_date)[:10] if r.reeval_date else "")
        return xlsx_response(wb, "risk_assessments.xlsx")
