"""Bulk question import for Final Exam Question Bank (CSV / XLSX) plus template generation."""
import io
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.ground_training.models import Subject, Module
from .final_models import FinalExamQuestion, FinalQuestionType, QuestionDifficulty

FIELDNAMES = ['subject_code', 'module_title', 'question_text', 'question_type', 'difficulty', 'options', 'correct_answer', 'explanation']
TYPE_MAP = {v.label.lower(): v.value for v in FinalQuestionType}
DIFF_MAP = {v.label.lower(): v.value for v in QuestionDifficulty}


def _parse_options(raw):
    if not raw or not str(raw).strip():
        return []
    return [o.strip() for o in str(raw).split('|') if o.strip()]


def _parse_type(raw):
    val = str(raw).strip().lower()
    if val in TYPE_MAP:
        return TYPE_MAP[val]
    choices_lower = {v.value: v.value for v in FinalQuestionType}
    if val in choices_lower:
        return val
    return FinalQuestionType.MCQ


def _parse_diff(raw):
    val = str(raw).strip().lower()
    if val in DIFF_MAP:
        return DIFF_MAP[val]
    if val in dict(QuestionDifficulty.choices):
        return val
    return QuestionDifficulty.MEDIUM


def import_questions(file):
    """Import questions from an uploaded CSV or XLSX file. Returns {total, created, skipped, errors}."""
    ext = file.name.rsplit('.', 1)[-1].lower() if file.name else 'csv'
    rows = []

    if ext in ('xls', 'xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = {}
            for i, h in enumerate(headers):
                if i < len(row):
                    d[h] = str(row[i]).strip() if row[i] is not None else ''
            rows.append(d)
        wb.close()
    else:
        reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
        for r in reader:
            rows.append({k.strip(): v for k, v in r.items()})

    total = len(rows)
    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=1):
        try:
            subject_code = (row.get('subject_code') or '').strip()
            module_title = (row.get('module_title') or row.get('module') or '').strip()
            question_text = (row.get('question_text') or '').strip()
            if not subject_code or not question_text:
                errors.append({'row': i + 1, 'message': 'Missing subject_code or question_text'})
                skipped += 1
                continue

            try:
                subject = Subject.objects.get(code=subject_code)
            except Subject.DoesNotExist:
                errors.append({'row': i + 1, 'message': f'Subject not found: {subject_code}'})
                skipped += 1
                continue

            module = None
            if module_title:
                try:
                    module = Module.objects.get(title=module_title)
                except Module.DoesNotExist:
                    pass

            FinalExamQuestion.objects.create(
                subject=subject,
                module=module,
                question_text=question_text,
                question_type=_parse_type(row.get('question_type', '')),
                difficulty=_parse_diff(row.get('difficulty', '')),
                options=_parse_options(row.get('options', '')),
                correct_answer=(row.get('correct_answer') or '').strip(),
                explanation=(row.get('explanation') or '').strip(),
            )
            created += 1
        except Exception as e:
            errors.append({'row': i + 1, 'message': str(e)})
            skipped += 1

    return {'total': total, 'created': created, 'skipped': skipped, 'errors': errors}


def generate_template(fmt='csv'):
    """Return template bytes for CSV or XLSX."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Final Exam Questions'
        cols = FIELDNAMES + ['subject_code', 'module_title', 'question_text', 'question_type', 'difficulty', 'options', 'correct_answer', 'explanation']
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        ws.cell(row=2, column=1, value='PPL-AL')
        ws.cell(row=2, column=2, value='Air Law')
        ws.cell(row=2, column=3, value='What does ICAO stand for?')
        ws.cell(row=2, column=4, value='Multiple Choice')
        ws.cell(row=2, column=5, value='Easy')
        ws.cell(row=2, column=6, value='International Civil Aviation Org|International Civil Air Org|Intl Commercial Aviation Org')
        ws.cell(row=2, column=7, value='International Civil Aviation Org')
        ws.cell(row=2, column=8, value='ICAO is a UN specialized agency')
        for col in range(1, len(cols) + 1):
            ws.column_dimensions[chr(64 + col)].width = 35
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(FIELDNAMES)
    writer.writerow(['PPL-AL', 'Air Law', 'What does ICAO stand for?', 'Multiple Choice', 'Easy',
                     'International Civil Aviation Org|International Civil Air Org', 'International Civil Aviation Org',
                     'ICAO is a UN specialized agency'])
    return out.getvalue().encode('utf-8-sig')
