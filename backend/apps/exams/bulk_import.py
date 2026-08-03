"""Bulk question import for the Question Bank (CSV / XLSX) plus template generation."""
import io
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.students.models import TrainingProgram
from apps.ground_training.models import Subject, Module
from .models import QuestionBank, QuestionType

# Canonical columns: (field, header)
COLUMNS = [
    ('question_text', 'question_text'),
    ('question_type', 'question_type'),
    ('difficulty', 'difficulty'),
    ('program', 'program'),
    ('subject_code', 'subject_code'),
    ('module', 'module'),
    ('options', 'options'),
    ('correct_answer', 'correct_answer'),
    ('explanation', 'explanation'),
    ('reference', 'reference'),
]

HEADER_LABELS = {
    'question_text': 'Question Text',
    'question_type': 'Question Type',
    'difficulty': 'Difficulty',
    'program': 'Program',
    'subject_code': 'Subject Code',
    'module': 'Module',
    'options': 'Options',
    'correct_answer': 'Correct Answer',
    'explanation': 'Explanation',
    'reference': 'Reference',
}

# Accept common human-readable aliases for question type
QUESTION_TYPE_ALIASES = {
    'mcq': 'mcq',
    'multiple choice': 'mcq',
    'multiple_choice': 'mcq',
    'single choice': 'mcq',
    'true/false': 'true_false',
    'true_false': 'true_false',
    'true false': 'true_false',
    'truefalse': 'true_false',
    'boolean': 'true_false',
    'short answer': 'short_answer',
    'short_answer': 'short_answer',
    'essay': 'essay',
    'matching': 'matching',
    'ordering': 'ordering',
    'order': 'ordering',
    'case study': 'case_study',
    'case_study': 'case_study',
}

DIFFICULTIES = {'easy', 'medium', 'hard'}
PROGRAM_VALUES = {v for v, _ in TrainingProgram.choices}

EXAMPLE_ROW = {
    'question_text': 'What is the correct answer to this sample question?',
    'question_type': 'mcq',
    'difficulty': 'easy',
    'program': 'PPL',
    'subject_code': '010',
    'module': 'Introduction',
    'options': 'Option A|Option B|Option C|Option D',
    'correct_answer': 'Option B',
    'explanation': 'Short explanation of why Option B is correct.',
    'reference': 'PPL Air Law Manual Ch. 1',
}


def _normalize(value):
    if value is None:
        return ''
    return str(value).strip()


def _header_key(raw):
    """Normalize a header cell to a canonical field key."""
    if raw is None:
        return None
    s = str(raw).strip().lower().replace(' ', '_').replace('-', '_')
    # Map English display headers to canonical keys
    display_to_key = {v.lower().replace(' ', '_'): k for k, v in HEADER_LABELS.items()}
    return display_to_key.get(s, s)


def _split_options(value):
    """Options may be separated by newlines or pipes."""
    raw = _normalize(value)
    if not raw:
        return []
    return [o.strip() for o in raw.replace('|', '\n').split('\n') if o.strip()]


def _normalize_question_type(value):
    key = _normalize(value).lower()
    return QUESTION_TYPE_ALIASES.get(key)


def _resolve_subject(subject_code, subject_title, subject_by_code, subject_by_title):
    code = _normalize(subject_code)
    title = _normalize(subject_title)
    if code:
        obj = subject_by_code.get(code.lower())
        if obj:
            return obj
        # Also try to interpret the subject cell as a title when it isn't a code
        obj = subject_by_title.get(code.lower())
        if obj:
            return obj
    if title:
        return subject_by_title.get(title.lower())
    return None


def _resolve_module(module_title, subject_obj, modules_by_title, modules_by_subject_title):
    title = _normalize(module_title)
    if not title:
        return None
    if subject_obj is not None:
        obj = modules_by_subject_title.get((subject_obj.id, title.lower()))
        if obj:
            return obj
    return modules_by_title.get(title.lower())


def _load_lookups():
    subjects = list(Subject.objects.all())
    modules = list(Module.objects.select_related('subject').all())
    subject_by_code = {s.code.lower(): s for s in subjects}
    subject_by_title = {s.title_en.lower(): s for s in subjects}
    modules_by_title = {}
    modules_by_subject_title = {}
    for m in modules:
        modules_by_title.setdefault(m.title.lower(), m)
        if m.subject_id:
            modules_by_subject_title.setdefault((m.subject_id, m.title.lower()), m)
    return subject_by_code, subject_by_title, modules_by_title, modules_by_subject_title


def _rows_from_xlsx(file_obj):
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _rows_from_csv(file_obj):
    raw = file_obj.read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')
    return list(csv.reader(io.StringIO(text)))


def _parse_rows(rows, subject_by_code, subject_by_title, modules_by_title, modules_by_subject_title):
    if not rows:
        return [], 'Empty file'
    # Normalize headers: match by canonical key or display label
    header = [_header_key(h) for h in rows[0]]
    if 'question_text' not in header:
        return [], 'Missing "question_text" column (expected header row)'
    records = []
    for idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        rec = {}
        for i, key in enumerate(header):
            if key and i < len(row):
                rec[key] = row[i]
        if not any(_normalize(v) for v in rec.values()):
            continue  # skip fully empty rows
        records.append((idx, rec))
    return records, None


def import_questions(file):
    """Import questions from an uploaded CSV/XLSX file.

    Returns a summary dict: {total, created, errors: [{row, message}]}.
    """
    name = (getattr(file, 'name', '') or '').lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        rows = _rows_from_xlsx(file)
    else:
        rows = _rows_from_csv(file)

    (subject_by_code, subject_by_title, modules_by_title,
     modules_by_subject_title) = _load_lookups()

    records, parse_error = _parse_rows(rows, subject_by_code, subject_by_title,
                                       modules_by_title, modules_by_subject_title)
    if parse_error:
        return {'total': 0, 'created': 0, 'skipped': 0, 'errors': [{'row': 1, 'message': parse_error}]}

    created = 0
    errors = []
    for row_no, rec in records:
        question_text = _normalize(rec.get('question_text'))
        if not question_text:
            errors.append({'row': row_no, 'message': 'Missing question_text'})
            continue

        qtype = _normalize_question_type(rec.get('question_type')) or QuestionType.MCQ

        difficulty = _normalize(rec.get('difficulty')).lower() or 'easy'
        if difficulty not in DIFFICULTIES:
            errors.append({'row': row_no, 'message': f'Invalid difficulty "{difficulty}" (easy/medium/hard)'})
            continue

        program = _normalize(rec.get('program')).upper() or None
        if program and program not in PROGRAM_VALUES:
            errors.append({'row': row_no, 'message': f'Invalid program "{program}"'})
            continue

        subject_obj = _resolve_subject(rec.get('subject_code'), rec.get('subject_title'),
                                       subject_by_code, subject_by_title)
        module_obj = _resolve_module(rec.get('module'), subject_obj,
                                     modules_by_title, modules_by_subject_title)
        if subject_obj is None and module_obj is not None:
            subject_obj = module_obj.subject

        correct_answer = _normalize(rec.get('correct_answer'))
        if not correct_answer:
            errors.append({'row': row_no, 'message': 'Missing correct_answer'})
            continue

        options = _split_options(rec.get('options'))
        if qtype in (QuestionType.MCQ, QuestionType.TRUE_FALSE) and not options:
            errors.append({'row': row_no, 'message': 'MCQ/True-False requires options'})
            continue

        try:
            QuestionBank.objects.create(
                subject=subject_obj,
                question_text=question_text,
                question_type=qtype,
                options=options,
                correct_answer=correct_answer,
                explanation=_normalize(rec.get('explanation')) or None,
                reference=_normalize(rec.get('reference')) or None,
                difficulty=difficulty,
                program=program,
                module=module_obj,
            )
            created += 1
        except Exception as exc:  # pragma: no cover - DB errors are surfaced as row errors
            errors.append({'row': row_no, 'message': str(exc)[:200]})

    return {
        'total': len(records),
        'created': created,
        'skipped': len(records) - created,
        'errors': errors,
    }


def generate_template(fmt='csv'):
    """Return template bytes for CSV or XLSX."""
    headers = [HEADER_LABELS[k] for k, _ in COLUMNS]
    example = [EXAMPLE_ROW[k] for k, _ in COLUMNS]

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Questions'
        ws.append(headers)
        ws.append(example)
        bold = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='C4943C')
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill
        widths = [50, 16, 12, 10, 14, 20, 40, 30, 40, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

        notes = wb.create_sheet('Allowed Values')
        notes.append(['Column', 'Allowed values'])
        notes.append(['question_type', 'mcq, true_false, short_answer, essay, matching, ordering, case_study'])
        notes.append(['difficulty', 'easy, medium, hard'])
        notes.append(['program', ', '.join(v for v, _ in TrainingProgram.choices)])
        notes.append(['options', 'Separate options with a new line or a pipe (|). Required for mcq / true_false.'])
        notes.append(['subject_code', 'A subject code or title already existing in the system (optional).'])
        notes.append(['module', 'A module title already existing in the system (optional).'])
        for cell in notes[1]:
            cell.font = bold
            cell.fill = fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(example)
    return buf.getvalue().encode('utf-8-sig')
